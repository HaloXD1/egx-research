from __future__ import annotations

import ast
import calendar
from dataclasses import dataclass
from datetime import UTC, datetime
from html import unescape
from io import BytesIO, StringIO
from pathlib import Path
import re
from typing import Tuple

import numpy as np
import pandas as pd
import requests
import urllib3
from openpyxl import load_workbook

from egx_research.fundamentals import REQUIRED_COLUMNS
from egx_research.stock_rotation_config import StockRotationConfig
from egx_research.utils import write_json


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


@dataclass
class StockUniverseRow:
    ticker: str
    ric_code: str
    holding_name: str
    weight: float
    stock_page_url: str
    historical_csv_url: str


def _get(url: str) -> requests.Response:
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, verify=False, timeout=30)
    response.raise_for_status()
    return response


def _search_page_url(config: StockRotationConfig, ticker: str, holding_name: str) -> str:
    response = requests.get(
        config.sources.mubasher_company_search_url,
        params={"query": holding_name},
        headers={"User-Agent": "Mozilla/5.0"},
        verify=False,
        timeout=30,
    )
    response.raise_for_status()
    results = response.json()
    if not isinstance(results, list):
        raise ValueError(f"Unexpected search response for {ticker}: {results}")

    preferred = None
    for item in results:
        url = item.get("url", "")
        label = item.get("label", "")
        if f"/markets/EGX/stocks/" not in url:
            continue
        if label.startswith(f"{ticker} -") or f" {ticker} " in f" {label} ":
            preferred = url
            break
        if preferred is None:
            preferred = url
    if preferred is None:
        raise ValueError(f"Could not resolve Mubasher page for {ticker}")
    return f"https://english.mubasher.info{preferred}"


def fetch_current_constituents(config: StockRotationConfig) -> Tuple[pd.DataFrame, pd.Timestamp]:
    content = _get(config.sources.etf_rebalancing_url).content
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook["Basket Constituents"]
    rows: list[StockUniverseRow] = []
    started = False
    last_update = pd.Timestamp.now(tz=UTC).normalize()

    for values in sheet.iter_rows(values_only=True):
        if not started:
            if values and str(values[0]).strip().lower().startswith("last update"):
                last_update = pd.Timestamp(values[1]).normalize()
            if values and values[0] == "RIC Code":
                started = True
            continue
        if not values or values[0] is None:
            break
        ric_code = str(values[0]).strip()
        ticker = ric_code.split(".")[0].upper()
        holding_name = str(values[1]).strip()
        weight = float(values[6]) if values[6] is not None else 0.0
        stock_page_url = config.sources.mubasher_stock_url_template.format(ticker=ticker)
        rows.append(
            StockUniverseRow(
                ticker=ticker,
                ric_code=ric_code,
                holding_name=holding_name,
                weight=weight,
                stock_page_url=stock_page_url,
                historical_csv_url="",
            )
        )

    return pd.DataFrame([row.__dict__ for row in rows]), last_update


def resolve_historical_csv_url(config: StockRotationConfig, ticker: str, holding_name: str, stock_page_url: str) -> tuple[str, str]:
    try:
        html = _get(stock_page_url).text
        resolved_page_url = stock_page_url
    except Exception:
        resolved_page_url = _search_page_url(config, ticker, holding_name)
        html = _get(resolved_page_url).text
    match = re.search(r'historical-data-url="([^"]+\.csv)"', html)
    if not match:
        raise ValueError(f"Could not find historical-data-url on page: {resolved_page_url}")
    return resolved_page_url, match.group(1)


def parse_mubasher_history_csv(csv_text: str) -> pd.DataFrame:
    frame = pd.read_csv(
        StringIO(csv_text),
        header=None,
        names=["date", "open", "high", "low", "close", "volume"],
    )
    frame["date"] = pd.to_datetime(frame["date"].str.split("/").str[0], errors="coerce")
    for column in ("open", "high", "low", "close", "volume"):
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame = frame.dropna(subset=["date", "open", "high", "low", "close"])
    frame = frame[(frame[["open", "high", "low", "close"]] > 0.0).all(axis=1)]
    frame = frame.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)
    frame["high"] = frame[["high", "open", "close"]].max(axis=1)
    frame["low"] = frame[["low", "open", "close"]].min(axis=1)
    frame["volume"] = frame["volume"].fillna(0.0)
    return frame


def sync_stock_rotation_data(config: StockRotationConfig) -> Path:
    root_dir = Path(config.storage.root_dir)
    raw_dir = root_dir / config.storage.raw_dir
    normalized_dir = root_dir / config.storage.normalized_dir
    raw_dir.mkdir(parents=True, exist_ok=True)
    normalized_dir.mkdir(parents=True, exist_ok=True)

    universe, effective_date = fetch_current_constituents(config)
    normalized_frames: list[pd.DataFrame] = []
    resolved_urls: list[str] = []

    for row in universe.itertuples(index=False):
        resolved_page_url, historical_csv_url = resolve_historical_csv_url(config, row.ticker, row.holding_name, row.stock_page_url)
        resolved_urls.append(historical_csv_url)
        csv_text = _get(historical_csv_url).text
        raw_path = raw_dir / f"{row.ticker}.csv"
        raw_path.write_text(csv_text, encoding="utf-8")

        normalized = parse_mubasher_history_csv(csv_text)
        normalized["symbol"] = row.ticker
        normalized["holding_name"] = row.holding_name
        normalized["weight"] = row.weight
        normalized_path = normalized_dir / f"{row.ticker}.csv"
        normalized.to_csv(normalized_path, index=False)
        normalized_frames.append(normalized)
        universe.loc[universe["ticker"] == row.ticker, "stock_page_url"] = resolved_page_url

    universe["historical_csv_url"] = resolved_urls
    universe.to_csv(root_dir / config.storage.universe_filename, index=False)

    panel = pd.concat(normalized_frames, ignore_index=True).sort_values(["date", "symbol"]).reset_index(drop=True)
    panel.to_csv(root_dir / config.storage.panel_filename, index=False)

    membership_path = root_dir / config.storage.membership_filename
    earliest_date = pd.Timestamp(panel["date"].min()).normalize()
    current_members = universe[["ticker"]].rename(columns={"ticker": "symbol"})
    current_members["is_member"] = True

    snapshots = []
    if membership_path.exists():
        snapshots = [pd.read_csv(membership_path, parse_dates=["effective_date"])]

    baseline = current_members.copy()
    baseline["effective_date"] = earliest_date
    baseline["source"] = "backfilled_current_constituents"
    current_snapshot = current_members.copy()
    current_snapshot["effective_date"] = effective_date
    current_snapshot["source"] = "official_etf_rebalancing_workbook"
    snapshots.extend([baseline, current_snapshot])

    membership = pd.concat(snapshots, ignore_index=True)
    membership["effective_date"] = pd.to_datetime(membership["effective_date"]).dt.normalize()
    membership = membership.sort_values(["effective_date", "symbol", "source"]).drop_duplicates(
        subset=["effective_date", "symbol"],
        keep="last",
    )
    membership.to_csv(membership_path, index=False)
    return root_dir


NEWS_CARD_PATTERN = re.compile(
    r'<span class="mi-article-media-block__date">(?P<date>.*?)</span>.*?'
    r'<a class="mi-article-media-block__title" href="(?P<href>/news/[^"]+)">(?P<title>.*?)</a>.*?'
    r'<div class="mi-hide-for-small mi-article-media-block__text"[^>]*>(?P<summary>.*?)</div>',
    flags=re.S,
)


def _strip_html(value: str) -> str:
    return re.sub(r"<.*?>", " ", unescape(value)).replace("\xa0", " ").strip()


def _parse_number(value: object) -> float | None:
    if value is None:
        return None
    text = _strip_html(str(value)).replace(",", "").strip()
    if text in {"", "-", "None", "none", "null"}:
        return None
    match = re.search(r"-?[0-9]+(?:\.[0-9]+)?(?:E[+-]?[0-9]+)?", text, flags=re.I)
    if not match:
        return None
    return float(match.group(0))


def _extract_stock_overview_number(html: str, label: str) -> float | None:
    marker = f'<span class="stock-overview__text">{label}</span>'
    index = html.find(marker)
    if index < 0:
        return None
    block = html[index : index + 900]
    match = re.search(
        r'<span class="(?:number number--aligned|stock-overview__value[^"]*)">([^<]+)</span>',
        block,
        flags=re.S,
    )
    if not match:
        match = re.search(r'<span class="number number--aligned">([^<]+)</span>', block)
    return _parse_number(match.group(1)) if match else None


def _extract_stock_overview_based_period(html: str, label: str) -> tuple[str, int] | None:
    marker = f'<span class="stock-overview__text">{label}</span>'
    index = html.find(marker)
    if index < 0:
        return None
    block = html[index : index + 1000]
    match = re.search(r"Based on:\s*([A-Za-z ]+)\s+([0-9]{4})", _strip_html(block))
    if not match:
        return None
    return match.group(1).strip(), int(match.group(2))


def parse_mubasher_stock_overview(html: str) -> dict[str, object]:
    eps_period = _extract_stock_overview_based_period(html, "EPS")
    bvps_period = _extract_stock_overview_based_period(html, "Book Value (BVPS)")
    return {
        "market_cap": _extract_stock_overview_number(html, "Market Cap"),
        "book_value_per_share": _extract_stock_overview_number(
            html, "Book Value (BVPS)"
        ),
        "book_value_period": bvps_period,
        "pb_ratio": _extract_stock_overview_number(html, "P/B Ratio"),
        "eps": _extract_stock_overview_number(html, "EPS"),
        "eps_period": eps_period,
        "pe_ratio": _extract_stock_overview_number(html, "P/E Ratio"),
        "shares_outstanding": _extract_stock_overview_number(
            html, "Current Total Shares"
        ),
    }


def _extract_financial_statement_payload(html: str) -> dict[str, object]:
    match = re.search(r"midata\.financialStatement\s*=\s*(\{.*?\});", html, flags=re.S)
    if not match:
        raise ValueError("Could not find Mubasher financialStatement payload.")
    payload = match.group(1).replace("null", "None")
    return ast.literal_eval(payload)


def _normalise_statement_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()


STATEMENT_ALIASES = {
    "revenue": [
        "total revenues",
        "revenue",
        "revenues",
        "sales",
        "net sales",
        "net operating revenue",
    ],
    "net_income": [
        "net income or loss",
        "net income",
        "net profit",
        "net profit loss",
        "profit loss",
    ],
    "equity": [
        "total owners equity minority interest equity",
        "total shareholders equity",
        "total stockholders equity",
        "total equity",
    ],
    "assets": ["total assets"],
    "operating_cf": [
        "net cash flow from used in operating activities",
        "net cash from operating activities",
        "cash flow from operating activities",
    ],
    "capex": [
        "capital expenditures",
        "purchase of property plant and equipment",
        "payments for property plant and equipment",
    ],
    "debt": ["total debt", "total liabilities"],
    "cash": ["cash and cash equivalents", "cash cash equivalents"],
}

STATEMENT_MONETARY_COLUMNS = [
    "revenue",
    "net_income",
    "equity",
    "assets",
    "operating_cf",
    "capex",
    "debt",
    "cash",
]


def _statement_records(period: dict[str, object]) -> dict[str, dict[str, object]]:
    records: dict[str, dict[str, object]] = {}
    for section in period.get("sections", []):
        if not isinstance(section, dict):
            continue
        for record in section.get("records", []):
            if not isinstance(record, dict):
                continue
            label = str(record.get("label", ""))
            values = record.get("values", {})
            if isinstance(values, dict):
                records[_normalise_statement_label(label)] = values
    return records


def _statement_value(
    records: dict[str, dict[str, object]], aliases: list[str], year: str
) -> float | None:
    for alias in aliases:
        values = records.get(_normalise_statement_label(alias))
        if values is None:
            continue
        value = _parse_number(values.get(str(year)))
        if value is not None:
            return value
    return None


def _period_quarter(label: str) -> int | None:
    normalized = label.lower()
    if "first quarter" in normalized:
        return 1
    if "second quarter" in normalized:
        return 2
    if "third quarter" in normalized:
        return 3
    if "fourth quarter" in normalized:
        return 4
    return None


def _period_end(year: int, label: str, year_start: str) -> pd.Timestamp:
    month_lookup = {name.lower(): index for index, name in enumerate(calendar.month_name) if name}
    start_month = month_lookup.get(str(year_start).strip().lower(), 1)
    quarter = _period_quarter(label)
    if quarter is None:
        end_offset = 11
    else:
        end_offset = quarter * 3 - 1
    zero_based = start_month - 1 + end_offset
    end_year = year + zero_based // 12
    end_month = zero_based % 12 + 1
    end_day = calendar.monthrange(end_year, end_month)[1]
    return pd.Timestamp(year=end_year, month=end_month, day=end_day)


def _fiscal_period(year: int, label: str) -> str:
    quarter = _period_quarter(label)
    if quarter is None:
        return f"FY {year}"
    return f"Q{quarter} {year}"


def _expand_two_digit_year(value: str) -> int:
    number = int(value)
    return 2000 + number if number < 100 else number


def _event_period_kind(text: str) -> str | None:
    lowered = text.lower()
    if re.search(r"\bq1\b|first quarter", lowered):
        return "Q1"
    if re.search(r"\bq2\b|second quarter|\bh1\b|first half", lowered):
        return "Q2"
    if re.search(r"\bq3\b|third quarter|\b9m\b|nine months", lowered):
        return "Q3"
    if re.search(r"\bq4\b|fourth quarter", lowered):
        return "Q4"
    if re.search(r"\bfy\b|fiscal year|financial year|annual", lowered):
        return "FY"
    return None


def _event_period_years(text: str) -> tuple[int | None, bool]:
    slash = re.search(r"(\d{2,4})\s*/\s*(\d{2,4})", text)
    if slash:
        return _expand_two_digit_year(slash.group(2)), True
    fiscal = re.search(r"\b(?:fy|fiscal year|financial year)[\s-]*(\d{2,4})\b", text, flags=re.I)
    if fiscal:
        return _expand_two_digit_year(fiscal.group(1)), False
    year = re.search(r"\b(20\d{2})\b", text)
    if year:
        return int(year.group(1)), False
    short = re.search(r"[-\s](\d{2})(?:\D|$)", text)
    if short:
        return _expand_two_digit_year(short.group(1)), False
    return None, False


def _period_end_from_event(kind: str, end_year: int, split_fiscal_year: bool) -> pd.Timestamp:
    if split_fiscal_year:
        mapping = {
            "Q1": (end_year - 1, 9, 30),
            "Q2": (end_year - 1, 12, 31),
            "Q3": (end_year, 3, 31),
            "Q4": (end_year, 6, 30),
            "FY": (end_year, 6, 30),
        }
    else:
        mapping = {
            "Q1": (end_year, 3, 31),
            "Q2": (end_year, 6, 30),
            "Q3": (end_year, 9, 30),
            "Q4": (end_year, 12, 31),
            "FY": (end_year, 12, 31),
        }
    year, month, day = mapping[kind]
    return pd.Timestamp(year=year, month=month, day=day)


def extract_disclosure_fiscal_period(
    title: str, summary: str, event_date: pd.Timestamp
) -> tuple[str, pd.Timestamp] | None:
    text = f"{title} {summary}"
    kind = _event_period_kind(text)
    if kind is None:
        return None
    year, split = _event_period_years(text)
    if year is None:
        year = pd.Timestamp(event_date).year
    fiscal_period = f"{kind} {year}"
    period_end = _period_end_from_event(kind, year, split)
    return fiscal_period, period_end


def _actual_disclosure_overrides(disclosure_events: pd.DataFrame) -> pd.DataFrame:
    if disclosure_events.empty:
        return pd.DataFrame(
            columns=["symbol", "fiscal_period", "actual_period_end", "actual_filing_date", "actual_source_url"]
        )
    rows: list[dict[str, object]] = []
    events = disclosure_events.copy()
    events["event_date"] = pd.to_datetime(events["event_date"], errors="coerce")
    events = events[events["event_class"].astype(str).str.lower().eq("earnings")]
    for row in events.itertuples(index=False):
        parsed = extract_disclosure_fiscal_period(
            str(getattr(row, "title", "")),
            str(getattr(row, "summary", "")),
            pd.Timestamp(getattr(row, "event_date")),
        )
        if parsed is None:
            continue
        fiscal_period, period_end = parsed
        event_date = pd.Timestamp(getattr(row, "event_date")).normalize()
        if event_date < period_end:
            continue
        rows.append(
            {
                "symbol": str(getattr(row, "symbol")).upper(),
                "fiscal_period": fiscal_period,
                "actual_period_end": period_end,
                "actual_filing_date": event_date,
                "actual_source_url": str(getattr(row, "source_url", "")),
            }
        )
    if not rows:
        return pd.DataFrame(
            columns=["symbol", "fiscal_period", "actual_period_end", "actual_filing_date", "actual_source_url"]
        )
    return (
        pd.DataFrame(rows)
        .sort_values(["symbol", "fiscal_period", "actual_filing_date"])
        .drop_duplicates(["symbol", "fiscal_period"], keep="first")
        .reset_index(drop=True)
    )


def apply_actual_disclosure_dates(
    fundamentals: pd.DataFrame, disclosure_events: pd.DataFrame
) -> tuple[pd.DataFrame, int]:
    overrides = _actual_disclosure_overrides(disclosure_events)
    if fundamentals.empty or overrides.empty:
        return fundamentals, 0
    frame = fundamentals.copy()
    merged = frame.merge(overrides, on=["symbol", "fiscal_period"], how="left")
    matched = merged["actual_filing_date"].notna()
    merged.loc[matched, "period_end"] = merged.loc[matched, "actual_period_end"]
    merged.loc[matched, "filing_date"] = merged.loc[matched, "actual_filing_date"]
    merged.loc[matched, "source_url"] = merged.loc[matched, "actual_source_url"].where(
        merged.loc[matched, "actual_source_url"].astype(str).ne(""),
        merged.loc[matched, "source_url"],
    )
    merged = merged.drop(
        columns=["actual_period_end", "actual_filing_date", "actual_source_url"]
    )
    return merged.reindex(columns=fundamentals.columns), int(matched.sum())


def filter_future_fundamentals(
    fundamentals: pd.DataFrame, as_of: pd.Timestamp | datetime
) -> tuple[pd.DataFrame, dict[str, object]]:
    as_of_date = pd.Timestamp(as_of)
    if as_of_date.tzinfo is not None:
        as_of_date = as_of_date.tz_convert(None)
    as_of_date = as_of_date.normalize()

    if fundamentals.empty:
        return fundamentals.copy(), {
            "fundamental_as_of_date": str(as_of_date.date()),
            "future_period_rows_removed": 0,
            "future_filing_rows_removed": 0,
            "future_rows_removed_total": 0,
        }

    frame = fundamentals.copy()
    period_end = pd.to_datetime(frame["period_end"], errors="coerce")
    filing_date = pd.to_datetime(frame["filing_date"], errors="coerce")
    future_period = period_end > as_of_date
    future_filing = filing_date > as_of_date
    future_any = future_period | future_filing
    filtered = frame.loc[~future_any].copy()
    return filtered.reindex(columns=fundamentals.columns), {
        "fundamental_as_of_date": str(as_of_date.date()),
        "future_period_rows_removed": int(future_period.sum()),
        "future_filing_rows_removed": int(future_filing.sum()),
        "future_rows_removed_total": int(future_any.sum()),
    }


def _currency_code(value: object) -> str:
    text = str(value or "").upper()
    match = re.search(r"\(([A-Z]{3})\)", text)
    if match:
        return match.group(1)
    if "EGYPTIAN POUND" in text:
        return "EGP"
    return str(value or "").strip()


def _period_source_url(
    period: dict[str, object], year: str, fallback_url: str
) -> str:
    attachments = period.get("attachments", {})
    if isinstance(attachments, dict):
        attachment = attachments.get(str(year))
        if attachment:
            return str(attachment)
    return fallback_url


def _unit_multiplier(
    values: dict[str, float | None],
    *,
    shares_outstanding: float | None,
    stats: dict[str, object],
    fiscal_period: str,
    year: int,
) -> float:
    if not shares_outstanding or shares_outstanding <= 0:
        return 1.0

    candidates = [1.0, 1_000.0, 1_000_000.0]
    score_inputs: list[tuple[float, float]] = []
    period_label = fiscal_period.split()[0]
    annualizer = 4.0 if period_label.startswith("Q") else 1.0

    eps_period = stats.get("eps_period")
    if eps_period is not None and isinstance(eps_period, tuple):
        eps_label, eps_year = eps_period
        eps = _parse_number(stats.get("eps"))
        net_income = values.get("net_income")
        if (
            eps is not None
            and net_income is not None
            and eps_year == year
            and _fiscal_period(year, eps_label).split()[0] == period_label
        ):
            score_inputs.append((net_income * annualizer / shares_outstanding, eps))

    bvps_period = stats.get("book_value_period")
    if bvps_period is not None and isinstance(bvps_period, tuple):
        bvps_label, bvps_year = bvps_period
        bvps = _parse_number(stats.get("book_value_per_share"))
        equity = values.get("equity")
        if (
            bvps is not None
            and equity is not None
            and bvps_year == year
            and _fiscal_period(year, bvps_label).split()[0] == period_label
        ):
            score_inputs.append((equity / shares_outstanding, bvps))

    best_multiplier = 1.0
    best_score = float("inf")
    for multiplier in candidates:
        score = 0.0
        for observed, target in score_inputs:
            if observed <= 0 or target <= 0:
                continue
            score += abs(np.log10((observed * multiplier) / target))
        equity = values.get("equity")
        if equity is not None and equity > 0:
            equity_ps = equity * multiplier / shares_outstanding
            if equity_ps < 0.05:
                score += 3.0
        assets = values.get("assets")
        market_cap = _parse_number(stats.get("market_cap"))
        if assets is not None and assets > 0 and market_cap is not None:
            price = market_cap / shares_outstanding
            assets_ps = assets * multiplier / shares_outstanding
            if price > 0 and assets_ps < price * 0.03:
                score += 2.0
        if score < best_score:
            best_score = score
            best_multiplier = multiplier
    return best_multiplier


def _smooth_statement_unit_outliers(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or "shares_outstanding" not in frame.columns:
        return frame

    result = frame.copy()
    shares = pd.to_numeric(result["shares_outstanding"], errors="coerce")
    shares_value = float(shares.dropna().iloc[0]) if shares.notna().any() else 0.0
    if shares_value <= 0:
        return result

    scale_rows = pd.Series(False, index=result.index)
    for column, min_reference in [("equity", 0.5), ("assets", 0.5)]:
        values = pd.to_numeric(result[column], errors="coerce")
        per_share = values / shares_value
        reference_values = per_share[per_share > min_reference]
        if reference_values.empty:
            continue
        reference = float(reference_values.median())
        tiny = (per_share > 0.0) & (per_share < reference / 100.0)
        scaled_back_near_reference = (per_share * 1000.0).between(
            reference / 10.0, reference * 10.0
        )
        scale_rows |= tiny & scaled_back_near_reference

    for column in STATEMENT_MONETARY_COLUMNS:
        if column in result.columns:
            result.loc[scale_rows, column] = (
                pd.to_numeric(result.loc[scale_rows, column], errors="coerce")
                * 1000.0
            )
    return result


def parse_mubasher_financial_statements(
    html: str,
    *,
    symbol: str,
    source_url: str,
    stock_overview: dict[str, object] | None = None,
    quarterly_lag_days: int = 60,
    annual_lag_days: int = 90,
) -> pd.DataFrame:
    statement = _extract_financial_statement_payload(html)
    stats = stock_overview or {}
    shares_outstanding = _parse_number(stats.get("shares_outstanding"))
    rows: list[dict[str, object]] = []
    currency = _currency_code(statement.get("currency"))
    year_start = str(statement.get("yearStart") or "January")

    for period in statement.get("periods", []):
        if not isinstance(period, dict):
            continue
        label = str(period.get("label", "")).strip()
        if not label:
            continue
        records = _statement_records(period)
        years = [str(year) for year in period.get("years", [])]
        for year_text in years:
            year = int(year_text)
            values = {
                column: _statement_value(records, aliases, year_text)
                for column, aliases in STATEMENT_ALIASES.items()
            }
            if all(values.get(column) is None for column in ["net_income", "equity", "assets"]):
                continue
            fiscal_period = _fiscal_period(year, label)
            multiplier = _unit_multiplier(
                values,
                shares_outstanding=shares_outstanding,
                stats=stats,
                fiscal_period=fiscal_period,
                year=year,
            )
            scaled_values = {
                column: (None if value is None else value * multiplier)
                for column, value in values.items()
            }
            end = _period_end(year, label, year_start)
            lag_days = annual_lag_days if fiscal_period.startswith("FY") else quarterly_lag_days
            rows.append(
                {
                    "symbol": symbol.upper(),
                    "period_end": end,
                    "filing_date": end + pd.Timedelta(days=int(lag_days)),
                    "fiscal_period": fiscal_period,
                    "currency": currency,
                    "source_url": _period_source_url(period, year_text, source_url),
                    "revenue": scaled_values["revenue"],
                    "net_income": scaled_values["net_income"],
                    "equity": scaled_values["equity"],
                    "assets": scaled_values["assets"],
                    "operating_cf": scaled_values["operating_cf"],
                    "capex": scaled_values["capex"],
                    "debt": scaled_values["debt"],
                    "cash": scaled_values["cash"],
                    "shares_outstanding": shares_outstanding,
                }
            )

    if not rows:
        return pd.DataFrame(columns=REQUIRED_COLUMNS)
    return (
        _smooth_statement_unit_outliers(pd.DataFrame(rows))
        .reindex(columns=REQUIRED_COLUMNS)
        .sort_values(["symbol", "period_end", "filing_date", "fiscal_period"])
        .reset_index(drop=True)
    )


def sync_stock_fundamentals(
    config: StockRotationConfig, *, as_of: pd.Timestamp | datetime | None = None
) -> Path:
    root_dir = Path(config.storage.root_dir)
    root_dir.mkdir(parents=True, exist_ok=True)
    as_of = as_of or datetime.now(UTC)
    universe_path = root_dir / config.storage.universe_filename
    if not universe_path.exists():
        raise FileNotFoundError(f"Universe missing: {universe_path}")

    universe = pd.read_csv(universe_path)
    frames: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []
    statement_path = config.sources.mubasher_financial_statements_path.strip("/")

    for row in universe.itertuples(index=False):
        symbol = str(getattr(row, "ticker", getattr(row, "symbol", ""))).upper()
        stock_page_url = str(
            getattr(
                row,
                "stock_page_url",
                config.sources.mubasher_stock_url_template.format(ticker=symbol),
            )
        )
        statements_url = f"{stock_page_url.rstrip('/')}/{statement_path}"
        try:
            stock_html = _get(stock_page_url).text
            overview = parse_mubasher_stock_overview(stock_html)
            statement_html = _get(statements_url).text
            fundamentals = parse_mubasher_financial_statements(
                statement_html,
                symbol=symbol,
                source_url=statements_url,
                stock_overview=overview,
                quarterly_lag_days=config.sources.fundamental_quarterly_lag_days,
                annual_lag_days=config.sources.fundamental_annual_lag_days,
            )
            frames.append(fundamentals)
            summary_rows.append(
                {
                    "symbol": symbol,
                    "stock_page_url": stock_page_url,
                    "statements_url": statements_url,
                    "rows_loaded": int(len(fundamentals)),
                    "shares_outstanding": overview.get("shares_outstanding"),
                    "status": "ok",
                }
            )
        except Exception as exc:
            summary_rows.append(
                {
                    "symbol": symbol,
                    "stock_page_url": stock_page_url,
                    "statements_url": statements_url,
                    "rows_loaded": 0,
                    "shares_outstanding": None,
                    "status": f"error:{exc}",
                }
            )

    fundamentals = (
        pd.concat(frames, ignore_index=True)
        if frames
        else pd.DataFrame(columns=REQUIRED_COLUMNS)
    )
    fundamentals = (
        fundamentals.reindex(columns=REQUIRED_COLUMNS)
        .sort_values(["symbol", "period_end", "filing_date", "fiscal_period"])
        .reset_index(drop=True)
    )
    disclosure_path = root_dir / config.storage.disclosure_events_filename
    actual_date_overrides = 0
    if disclosure_path.exists():
        disclosure_events = pd.read_csv(disclosure_path)
        fundamentals, actual_date_overrides = apply_actual_disclosure_dates(
            fundamentals, disclosure_events
        )
        fundamentals = (
            fundamentals.reindex(columns=REQUIRED_COLUMNS)
            .sort_values(["symbol", "period_end", "filing_date", "fiscal_period"])
            .reset_index(drop=True)
        )
    fundamentals, future_filter_stats = filter_future_fundamentals(
        fundamentals, as_of
    )
    fundamentals = (
        fundamentals.reindex(columns=REQUIRED_COLUMNS)
        .sort_values(["symbol", "period_end", "filing_date", "fiscal_period"])
        .reset_index(drop=True)
    )
    fundamentals.to_csv(root_dir / config.storage.fundamentals_filename, index=False)

    write_json(
        root_dir / "fundamental_sync_summary.json",
        {
            "created_at": datetime.now(UTC).isoformat(),
            "symbols_total": int(len(universe)),
            "symbols_with_rows": int(fundamentals["symbol"].nunique())
            if not fundamentals.empty
            else 0,
            "fundamental_rows_loaded": int(len(fundamentals)),
            "quarterly_lag_days": int(config.sources.fundamental_quarterly_lag_days),
            "annual_lag_days": int(config.sources.fundamental_annual_lag_days),
            "actual_disclosure_date_overrides": int(actual_date_overrides),
            **future_filter_stats,
            "source": "mubasher_financial_statements",
            "symbol_status": summary_rows,
        },
    )
    return root_dir


def _parse_mubasher_news_date(value: str) -> pd.Timestamp | None:
    cleaned = _strip_html(value)
    cleaned = re.sub(r"\b00:(\d{2}\s*[AP]M)\b", r"12:\1", cleaned, flags=re.I)
    current_year = pd.Timestamp.now(tz=UTC).year
    candidates = [
        (cleaned, "%d %B %Y %I:%M %p"),
        (cleaned, "%d %b %Y %I:%M %p"),
        (f"{cleaned} {current_year}", "%d %B %I:%M %p %Y"),
        (f"{cleaned} {current_year}", "%d %b %I:%M %p %Y"),
    ]
    for candidate, fmt in candidates:
        try:
            parsed = pd.to_datetime(candidate, format=fmt, errors="raise")
            return pd.Timestamp(parsed)
        except Exception:
            continue
    parsed = pd.to_datetime(cleaned, errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed)


def classify_disclosure_event(title: str, summary: str) -> str:
    text = f"{title} {summary}".lower()
    if any(token in text for token in ["dividend", "cash distribution", "cash dividends"]):
        return "dividend"
    if any(token in text for token in ["capital hike", "capital increase", "rights issue", "rights offering"]):
        return "capital_action"
    if any(token in text for token in ["agm", "egm", "minutes", "resolutions"]):
        return "governance"
    if any(token in text for token in ["financial statements", "results", "profit", "earnings"]):
        return "earnings"
    if any(token in text for token in ["contract", "award", "expansion", "investment", "acquisition", "due diligence"]):
        return "business_update"
    return "general"


def extract_dividend_cash_amount(summary: str) -> float | None:
    patterns = [
        r"EGP\s*([0-9]+(?:\.[0-9]+)?)\s*per share",
        r"cash dividends worth EGP\s*([0-9]+(?:\.[0-9]+)?)",
        r"dividends worth EGP\s*([0-9]+(?:\.[0-9]+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, summary, flags=re.I)
        if match:
            return float(match.group(1))
    return None


def parse_mubasher_news_page(
    html: str,
    *,
    symbol: str,
    source_url: str,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for match in NEWS_CARD_PATTERN.finditer(html):
        date_text = _strip_html(match.group("date"))
        title = _strip_html(match.group("title"))
        summary = _strip_html(match.group("summary"))
        href = match.group("href").strip()
        event_date = _parse_mubasher_news_date(date_text)
        if event_date is None:
            continue
        rows.append(
            {
                "symbol": symbol,
                "event_date": event_date.normalize(),
                "event_class": classify_disclosure_event(title, summary),
                "title": title,
                "summary": summary,
                "source_url": f"https://english.mubasher.info{href}",
                "source": "mubasher_news",
                "source_page_url": source_url,
                "cash_amount": extract_dividend_cash_amount(summary),
            }
        )
    if not rows:
        return pd.DataFrame(
            columns=[
                "symbol",
                "event_date",
                "event_class",
                "title",
                "summary",
                "source_url",
                "source",
                "source_page_url",
                "cash_amount",
            ]
        )
    frame = pd.DataFrame(rows)
    frame = (
        frame.sort_values(["symbol", "event_date", "source_url"])
        .drop_duplicates(subset=["symbol", "event_date", "title"], keep="last")
        .reset_index(drop=True)
    )
    return frame


def build_stage2_datasets_from_disclosures(events: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if events.empty:
        return (
            pd.DataFrame(columns=["symbol", "event_date", "cash_amount", "currency", "source"]),
            pd.DataFrame(columns=["symbol", "event_date", "event_type", "is_dilutive", "source"]),
        )

    dividends = events[events["event_class"] == "dividend"].copy()
    dividends = dividends[
        ["symbol", "event_date", "cash_amount", "source"]
    ].copy()
    dividends["currency"] = "EGP"
    dividends = dividends.sort_values(["symbol", "event_date"]).reset_index(drop=True)

    corp = events[events["event_class"].isin(["capital_action", "governance"])].copy()
    corp["event_type"] = corp["event_class"].map(
        {"capital_action": "capital_action", "governance": "governance"}
    )
    corp["is_dilutive"] = corp["title"].str.contains(
        r"capital hike|capital increase|rights issue|rights offering",
        flags=re.I,
        regex=True,
    )
    corp = corp[["symbol", "event_date", "event_type", "is_dilutive", "source"]].copy()
    corp = corp.sort_values(["symbol", "event_date"]).reset_index(drop=True)
    return dividends, corp


def sync_institutional_stage2_data(
    config: StockRotationConfig,
    *,
    max_news_items_per_symbol: int = 50,
) -> Path:
    root_dir = Path(config.storage.root_dir)
    root_dir.mkdir(parents=True, exist_ok=True)
    universe_path = root_dir / config.storage.universe_filename
    if not universe_path.exists():
        raise FileNotFoundError(f"Universe missing: {universe_path}")

    universe = pd.read_csv(universe_path)
    disclosure_frames: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []

    for row in universe.itertuples(index=False):
        stock_page_url = str(getattr(row, "stock_page_url"))
        symbol = str(getattr(row, "ticker"))
        news_url = stock_page_url.rstrip("/") + "/news"
        try:
            html = _get(news_url).text
            events = parse_mubasher_news_page(html, symbol=symbol, source_url=news_url)
            if max_news_items_per_symbol > 0 and not events.empty:
                events = (
                    events.sort_values("event_date", ascending=False)
                    .head(max_news_items_per_symbol)
                    .reset_index(drop=True)
                )
            disclosure_frames.append(events)
            summary_rows.append(
                {
                    "symbol": symbol,
                    "news_url": news_url,
                    "events_loaded": int(len(events)),
                    "status": "ok",
                }
            )
        except Exception as exc:
            summary_rows.append(
                {
                    "symbol": symbol,
                    "news_url": news_url,
                    "events_loaded": 0,
                    "status": f"error:{exc}",
                }
            )

    disclosure_events = (
        pd.concat(disclosure_frames, ignore_index=True)
        if disclosure_frames
        else pd.DataFrame(
            columns=[
                "symbol",
                "event_date",
                "event_class",
                "title",
                "summary",
                "source_url",
                "source",
                "source_page_url",
                "cash_amount",
            ]
        )
    )
    disclosure_events = disclosure_events.sort_values(
        ["symbol", "event_date", "source_url"], ascending=[True, False, True]
    ).reset_index(drop=True)

    dividend_actions, corporate_actions = build_stage2_datasets_from_disclosures(
        disclosure_events
    )

    disclosure_path = root_dir / config.storage.disclosure_events_filename
    dividend_path = root_dir / config.storage.dividend_actions_filename
    corporate_path = root_dir / config.storage.corporate_actions_filename

    disclosure_events.to_csv(disclosure_path, index=False)
    dividend_actions.to_csv(dividend_path, index=False)
    corporate_actions.to_csv(corporate_path, index=False)

    write_json(
        root_dir / "institutional_stage2_summary.json",
        {
            "created_at": datetime.now(UTC).isoformat(),
            "symbols_total": int(len(universe)),
            "symbols_with_events": int(
                pd.Series(disclosure_events["symbol"]).nunique()
                if not disclosure_events.empty
                else 0
            ),
            "disclosure_events_loaded": int(len(disclosure_events)),
            "dividend_actions_loaded": int(len(dividend_actions)),
            "corporate_actions_loaded": int(len(corporate_actions)),
            "max_news_items_per_symbol": int(max_news_items_per_symbol),
            "symbol_status": summary_rows,
        },
    )
    return root_dir
